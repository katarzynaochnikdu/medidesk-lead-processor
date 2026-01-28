"""
Porownanie wariantow: strukturalny vs surowy tekst
Oblicza wszystkie metryki z planu testow.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def load_data():
    """Wczytuje wszystkie pliki wynikowe."""
    print("==> Wczytuje dane...")
    
    # Referencja
    df_ref = pd.read_excel("test data reference result.xlsx")
    print(f"  Referencja: {len(df_ref)} wierszy")
    
    # Wariant 1 - strukturalny
    df_v1 = pd.read_excel("test_results_variant1.xlsx")
    print(f"  Wariant 1 (strukturalny): {len(df_v1)} wierszy")
    
    # Wariant 2 - surowy tekst
    df_v2 = pd.read_excel("test_results_variant2.xlsx")
    print(f"  Wariant 2 (surowy tekst): {len(df_v2)} wierszy")
    
    return df_ref, df_v1, df_v2


def get_ref_columns(df_ref):
    """Mapuje kolumny referencji (polskie znaki moga byc zle kodowane)."""
    # Uzywamy indeksow
    return {
        "contact_id": df_ref.columns[7],  # Powiazanie Kontaktu z bazy.id
        "contact_name": df_ref.columns[8],  # Powiazanie Kontaktu z bazy
        "account_id": df_ref.columns[9],  # Powiazanie Firmy z bazy.id
        "account_name": df_ref.columns[10],  # Powiazanie Firmy z bazy
    }


def calculate_contact_metrics(df_ref, df_test, ref_cols, variant_name):
    """Oblicza metryki identyfikacji kontaktow."""
    metrics = {
        "variant": variant_name,
        "true_positive": 0,
        "false_negative": 0,
        "false_positive": 0,
        "tier_4": 0,
        "tier_3": 0,
        "tier_2": 0,
        "tier_0": 0,
        "signal_E": 0,
        "signal_P": 0,
        "signal_L": 0,
        "signal_F": 0,
        "signal_A": 0,
    }
    
    details = []
    
    for idx in range(len(df_ref)):
        ref_contact_id = df_ref.iloc[idx][ref_cols["contact_id"]]
        ref_contact_name = df_ref.iloc[idx][ref_cols["contact_name"]]
        
        test_contact_id = df_test.iloc[idx].get("contact_id", "")
        test_contact_exists = df_test.iloc[idx].get("contact_exists", "NIE")
        test_contact_tier = df_test.iloc[idx].get("contact_tier", 0)
        test_contact_signals = df_test.iloc[idx].get("contact_signals", "")
        test_contact_name = df_test.iloc[idx].get("contact_name", "")
        
        ref_exists = pd.notna(ref_contact_id) and str(ref_contact_id).lower() != "nowy"
        test_exists = test_contact_exists == "TAK"
        
        detail = {
            "row": idx + 1,
            "ref_id": ref_contact_id if ref_exists else "NOWY",
            "ref_name": ref_contact_name if pd.notna(ref_contact_name) else "",
            "test_id": test_contact_id if test_exists else "",
            "test_name": test_contact_name,
            "test_tier": test_contact_tier,
            "test_signals": test_contact_signals,
            "result": "",
        }
        
        if ref_exists and test_exists:
            # Sprawdz czy to ten sam kontakt
            if str(ref_contact_id) == str(test_contact_id):
                metrics["true_positive"] += 1
                detail["result"] = "TRUE_POSITIVE"
            else:
                metrics["false_positive"] += 1
                detail["result"] = "FALSE_POSITIVE"
        elif ref_exists and not test_exists:
            metrics["false_negative"] += 1
            detail["result"] = "FALSE_NEGATIVE"
        elif not ref_exists and test_exists:
            metrics["false_positive"] += 1
            detail["result"] = "FALSE_POSITIVE"
        else:
            metrics["true_positive"] += 1  # Oba "nie istnieje" - poprawnie
            detail["result"] = "TRUE_NEGATIVE"
        
        # Tier distribution
        if test_contact_tier == 4:
            metrics["tier_4"] += 1
        elif test_contact_tier == 3:
            metrics["tier_3"] += 1
        elif test_contact_tier == 2:
            metrics["tier_2"] += 1
        else:
            metrics["tier_0"] += 1
        
        # Sygnaly
        if test_contact_signals and pd.notna(test_contact_signals) and isinstance(test_contact_signals, str):
            if "E" in test_contact_signals:
                metrics["signal_E"] += 1
            if "P" in test_contact_signals:
                metrics["signal_P"] += 1
            if "L" in test_contact_signals:
                metrics["signal_L"] += 1
            if "F" in test_contact_signals:
                metrics["signal_F"] += 1
            if "A" in test_contact_signals:
                metrics["signal_A"] += 1
        
        details.append(detail)
    
    return metrics, details


def calculate_account_metrics(df_ref, df_test, ref_cols, variant_name):
    """Oblicza metryki identyfikacji firm."""
    metrics = {
        "variant": variant_name,
        "true_positive": 0,
        "false_negative": 0,
        "false_positive": 0,
        "parent_correct": 0,
    }
    
    details = []
    
    for idx in range(len(df_ref)):
        ref_account_id = df_ref.iloc[idx][ref_cols["account_id"]]
        ref_account_name = df_ref.iloc[idx][ref_cols["account_name"]]
        
        test_account_id = df_test.iloc[idx].get("account_id", "")
        test_account_exists = df_test.iloc[idx].get("account_exists", "NIE")
        test_account_name = df_test.iloc[idx].get("account_name", "")
        test_account_match = df_test.iloc[idx].get("account_match", "")
        
        ref_exists = pd.notna(ref_account_id)
        test_exists = test_account_exists == "TAK"
        
        detail = {
            "row": idx + 1,
            "ref_id": ref_account_id if ref_exists else "",
            "ref_name": ref_account_name if pd.notna(ref_account_name) else "",
            "test_id": test_account_id if test_exists else "",
            "test_name": test_account_name,
            "test_match": test_account_match,
            "result": "",
        }
        
        if ref_exists and test_exists:
            if str(ref_account_id) == str(test_account_id):
                metrics["true_positive"] += 1
                metrics["parent_correct"] += 1
                detail["result"] = "TRUE_POSITIVE"
            else:
                metrics["false_positive"] += 1
                detail["result"] = "FALSE_POSITIVE"
        elif ref_exists and not test_exists:
            metrics["false_negative"] += 1
            detail["result"] = "FALSE_NEGATIVE"
        elif not ref_exists and test_exists:
            metrics["false_positive"] += 1
            detail["result"] = "FALSE_POSITIVE"
        else:
            metrics["true_positive"] += 1
            detail["result"] = "TRUE_NEGATIVE"
        
        details.append(detail)
    
    return metrics, details


def calculate_normalization_metrics(df_test, variant_name):
    """Oblicza metryki normalizacji danych."""
    metrics = {
        "variant": variant_name,
        "first_name_filled": 0,
        "last_name_filled": 0,
        "email_filled": 0,
        "phone_filled": 0,
        "company_filled": 0,
        "nip_found": 0,
        "nip_valid": 0,
        "nip_from_brave": 0,
    }
    
    for idx in range(len(df_test)):
        row = df_test.iloc[idx]
        
        if row.get("norm_first_name"):
            metrics["first_name_filled"] += 1
        if row.get("norm_last_name"):
            metrics["last_name_filled"] += 1
        if row.get("norm_email"):
            metrics["email_filled"] += 1
        if row.get("norm_phone"):
            metrics["phone_filled"] += 1
        if row.get("norm_company_name"):
            metrics["company_filled"] += 1
        if row.get("norm_nip"):
            metrics["nip_found"] += 1
        if row.get("norm_nip_valid") == "TAK":
            metrics["nip_valid"] += 1
        if row.get("norm_nip_source") == "Brave Search":
            metrics["nip_from_brave"] += 1
    
    return metrics


def calculate_recommendation_metrics(df_test, variant_name):
    """Oblicza metryki rekomendacji."""
    metrics = {
        "variant": variant_name,
        "create_new": 0,
        "link_to_existing": 0,
        "review_required": 0,
        "avg_confidence": 0.0,
    }
    
    confidences = []
    
    for idx in range(len(df_test)):
        row = df_test.iloc[idx]
        action = row.get("recommendation_action", "")
        
        if action == "create_new":
            metrics["create_new"] += 1
        elif action == "link_to_existing":
            metrics["link_to_existing"] += 1
        elif action == "review_required":
            metrics["review_required"] += 1
        
        conf = row.get("recommendation_confidence", 0)
        if pd.notna(conf):
            confidences.append(float(conf))
    
    if confidences:
        metrics["avg_confidence"] = sum(confidences) / len(confidences)
    
    return metrics


def create_comparison_report(df_ref, df_v1, df_v2, ref_cols):
    """Tworzy pelny raport porownawczy."""
    print("\n==> Obliczam metryki...")
    
    # Metryki kontaktow
    contact_v1, contact_details_v1 = calculate_contact_metrics(df_ref, df_v1, ref_cols, "STRUCTURAL")
    contact_v2, contact_details_v2 = calculate_contact_metrics(df_ref, df_v2, ref_cols, "RAW_TEXT")
    
    # Metryki firm
    account_v1, account_details_v1 = calculate_account_metrics(df_ref, df_v1, ref_cols, "STRUCTURAL")
    account_v2, account_details_v2 = calculate_account_metrics(df_ref, df_v2, ref_cols, "RAW_TEXT")
    
    # Metryki normalizacji
    norm_v1 = calculate_normalization_metrics(df_v1, "STRUCTURAL")
    norm_v2 = calculate_normalization_metrics(df_v2, "RAW_TEXT")
    
    # Metryki rekomendacji
    rec_v1 = calculate_recommendation_metrics(df_v1, "STRUCTURAL")
    rec_v2 = calculate_recommendation_metrics(df_v2, "RAW_TEXT")
    
    return {
        "contact": {"v1": contact_v1, "v2": contact_v2},
        "contact_details": {"v1": contact_details_v1, "v2": contact_details_v2},
        "account": {"v1": account_v1, "v2": account_v2},
        "account_details": {"v1": account_details_v1, "v2": account_details_v2},
        "normalization": {"v1": norm_v1, "v2": norm_v2},
        "recommendation": {"v1": rec_v1, "v2": rec_v2},
    }


def write_excel_report(metrics, df_ref, df_v1, df_v2, ref_cols, output_file):
    """Zapisuje raport do Excel."""
    print(f"\n==> Zapisuje raport do {output_file}...")
    
    wb = Workbook()
    
    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    winner_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loser_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === ARKUSZ 1: SUMMARY ===
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Tytul
    ws_summary["A1"] = "POROWNANIE WARIANTOW AI"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:D1")
    
    ws_summary["A2"] = "Wariant 1: STRUKTURALNY (osobne pola)"
    ws_summary["A3"] = "Wariant 2: SUROWY TEKST (wszystko w raw_name)"
    
    # Kontakty
    row = 5
    ws_summary[f"A{row}"] = "IDENTYFIKACJA KONTAKTOW"
    ws_summary[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    headers = ["Metryka", "Wariant 1 (STRUCT)", "Wariant 2 (RAW)", "Lepszy"]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    contact_rows = [
        ("True Positive", metrics["contact"]["v1"]["true_positive"], metrics["contact"]["v2"]["true_positive"], "max"),
        ("False Negative", metrics["contact"]["v1"]["false_negative"], metrics["contact"]["v2"]["false_negative"], "min"),
        ("False Positive", metrics["contact"]["v1"]["false_positive"], metrics["contact"]["v2"]["false_positive"], "min"),
        ("Tier 4", metrics["contact"]["v1"]["tier_4"], metrics["contact"]["v2"]["tier_4"], "max"),
        ("Tier 3", metrics["contact"]["v1"]["tier_3"], metrics["contact"]["v2"]["tier_3"], "max"),
        ("Tier 2", metrics["contact"]["v1"]["tier_2"], metrics["contact"]["v2"]["tier_2"], "info"),
        ("Sygnaly E", metrics["contact"]["v1"]["signal_E"], metrics["contact"]["v2"]["signal_E"], "max"),
        ("Sygnaly P", metrics["contact"]["v1"]["signal_P"], metrics["contact"]["v2"]["signal_P"], "max"),
        ("Sygnaly L", metrics["contact"]["v1"]["signal_L"], metrics["contact"]["v2"]["signal_L"], "max"),
        ("Sygnaly F", metrics["contact"]["v1"]["signal_F"], metrics["contact"]["v2"]["signal_F"], "max"),
        ("Sygnaly A", metrics["contact"]["v1"]["signal_A"], metrics["contact"]["v2"]["signal_A"], "max"),
    ]
    
    for label, v1, v2, compare in contact_rows:
        row += 1
        ws_summary.cell(row=row, column=1, value=label).border = thin_border
        cell_v1 = ws_summary.cell(row=row, column=2, value=v1)
        cell_v2 = ws_summary.cell(row=row, column=3, value=v2)
        cell_v1.border = thin_border
        cell_v2.border = thin_border
        
        if compare == "max":
            winner = "V1" if v1 > v2 else ("V2" if v2 > v1 else "REMIS")
            if v1 > v2:
                cell_v1.fill = winner_fill
            elif v2 > v1:
                cell_v2.fill = winner_fill
        elif compare == "min":
            winner = "V1" if v1 < v2 else ("V2" if v2 < v1 else "REMIS")
            if v1 < v2:
                cell_v1.fill = winner_fill
            elif v2 < v1:
                cell_v2.fill = winner_fill
        else:
            winner = "-"
        
        ws_summary.cell(row=row, column=4, value=winner).border = thin_border
    
    # Firmy
    row += 2
    ws_summary[f"A{row}"] = "IDENTYFIKACJA FIRM"
    ws_summary[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    account_rows = [
        ("True Positive", metrics["account"]["v1"]["true_positive"], metrics["account"]["v2"]["true_positive"], "max"),
        ("False Negative", metrics["account"]["v1"]["false_negative"], metrics["account"]["v2"]["false_negative"], "min"),
        ("False Positive", metrics["account"]["v1"]["false_positive"], metrics["account"]["v2"]["false_positive"], "min"),
        ("Parent Correct", metrics["account"]["v1"]["parent_correct"], metrics["account"]["v2"]["parent_correct"], "max"),
    ]
    
    for label, v1, v2, compare in account_rows:
        row += 1
        ws_summary.cell(row=row, column=1, value=label).border = thin_border
        cell_v1 = ws_summary.cell(row=row, column=2, value=v1)
        cell_v2 = ws_summary.cell(row=row, column=3, value=v2)
        cell_v1.border = thin_border
        cell_v2.border = thin_border
        
        if compare == "max":
            winner = "V1" if v1 > v2 else ("V2" if v2 > v1 else "REMIS")
            if v1 > v2:
                cell_v1.fill = winner_fill
            elif v2 > v1:
                cell_v2.fill = winner_fill
        elif compare == "min":
            winner = "V1" if v1 < v2 else ("V2" if v2 < v1 else "REMIS")
            if v1 < v2:
                cell_v1.fill = winner_fill
            elif v2 < v1:
                cell_v2.fill = winner_fill
        else:
            winner = "-"
        
        ws_summary.cell(row=row, column=4, value=winner).border = thin_border
    
    # Normalizacja
    row += 2
    ws_summary[f"A{row}"] = "NORMALIZACJA DANYCH"
    ws_summary[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    norm_rows = [
        ("Imie wypelnione", metrics["normalization"]["v1"]["first_name_filled"], metrics["normalization"]["v2"]["first_name_filled"], "max"),
        ("Nazwisko wypelnione", metrics["normalization"]["v1"]["last_name_filled"], metrics["normalization"]["v2"]["last_name_filled"], "max"),
        ("Email wypelniony", metrics["normalization"]["v1"]["email_filled"], metrics["normalization"]["v2"]["email_filled"], "max"),
        ("Telefon wypelniony", metrics["normalization"]["v1"]["phone_filled"], metrics["normalization"]["v2"]["phone_filled"], "max"),
        ("Firma wypelniona", metrics["normalization"]["v1"]["company_filled"], metrics["normalization"]["v2"]["company_filled"], "max"),
        ("NIP znaleziony", metrics["normalization"]["v1"]["nip_found"], metrics["normalization"]["v2"]["nip_found"], "max"),
        ("NIP poprawny", metrics["normalization"]["v1"]["nip_valid"], metrics["normalization"]["v2"]["nip_valid"], "max"),
        ("NIP z Brave", metrics["normalization"]["v1"]["nip_from_brave"], metrics["normalization"]["v2"]["nip_from_brave"], "info"),
    ]
    
    for label, v1, v2, compare in norm_rows:
        row += 1
        ws_summary.cell(row=row, column=1, value=label).border = thin_border
        cell_v1 = ws_summary.cell(row=row, column=2, value=v1)
        cell_v2 = ws_summary.cell(row=row, column=3, value=v2)
        cell_v1.border = thin_border
        cell_v2.border = thin_border
        
        if compare == "max":
            winner = "V1" if v1 > v2 else ("V2" if v2 > v1 else "REMIS")
            if v1 > v2:
                cell_v1.fill = winner_fill
            elif v2 > v1:
                cell_v2.fill = winner_fill
        else:
            winner = "-"
        
        ws_summary.cell(row=row, column=4, value=winner).border = thin_border
    
    # Rekomendacje
    row += 2
    ws_summary[f"A{row}"] = "REKOMENDACJE"
    ws_summary[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    rec_rows = [
        ("create_new", metrics["recommendation"]["v1"]["create_new"], metrics["recommendation"]["v2"]["create_new"], "info"),
        ("link_to_existing", metrics["recommendation"]["v1"]["link_to_existing"], metrics["recommendation"]["v2"]["link_to_existing"], "max"),
        ("review_required", metrics["recommendation"]["v1"]["review_required"], metrics["recommendation"]["v2"]["review_required"], "min"),
        ("Avg Confidence", f"{metrics['recommendation']['v1']['avg_confidence']:.0%}", f"{metrics['recommendation']['v2']['avg_confidence']:.0%}", "max"),
    ]
    
    for label, v1, v2, compare in rec_rows:
        row += 1
        ws_summary.cell(row=row, column=1, value=label).border = thin_border
        cell_v1 = ws_summary.cell(row=row, column=2, value=v1)
        cell_v2 = ws_summary.cell(row=row, column=3, value=v2)
        cell_v1.border = thin_border
        cell_v2.border = thin_border
        ws_summary.cell(row=row, column=4, value="-").border = thin_border
    
    # Kolumny
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 22
    ws_summary.column_dimensions['C'].width = 22
    ws_summary.column_dimensions['D'].width = 12
    
    # === ARKUSZ 2: FULL DETAILS - pelna dokumentacja kazdego rekordu ===
    ws_full = wb.create_sheet("Full Details")
    
    full_headers = [
        "Row",
        # Dane wejsciowe (oryginalne)
        "INPUT raw_name", "INPUT first_name", "INPUT last_name", "INPUT email", "INPUT phone", "INPUT company",
        # V1 - po normalizacji AI
        "V1 first_name", "V1 last_name", "V1 email", "V1 phone", "V1 company", "V1 NIP", "V1 NIP valid",
        # V1 - weryfikacja Zoho
        "V1 Contact exists", "V1 Contact ID", "V1 Contact name", "V1 Tier", "V1 Signals",
        "V1 Account exists", "V1 Account ID", "V1 Account name", "V1 Match reason",
        "V1 Recommendation", "V1 Confidence",
        # V2 - po normalizacji AI
        "V2 first_name", "V2 last_name", "V2 email", "V2 phone", "V2 company", "V2 NIP", "V2 NIP valid",
        # V2 - weryfikacja Zoho
        "V2 Contact exists", "V2 Contact ID", "V2 Contact name", "V2 Tier", "V2 Signals",
        "V2 Account exists", "V2 Account ID", "V2 Account name", "V2 Match reason",
        "V2 Recommendation", "V2 Confidence",
        # Referencja
        "REF Contact ID", "REF Contact name", "REF Account ID", "REF Account name",
        # Wynik porownania
        "Contact V1 Result", "Contact V2 Result", "Account V1 Result", "Account V2 Result",
    ]
    
    for col, h in enumerate(full_headers, 1):
        cell = ws_full.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    contact_details_v1 = metrics["contact_details"]["v1"]
    contact_details_v2 = metrics["contact_details"]["v2"]
    account_details_v1 = metrics["account_details"]["v1"]
    account_details_v2 = metrics["account_details"]["v2"]
    
    for i in range(len(df_v1)):
        row = i + 2
        v1 = df_v1.iloc[i]
        v2 = df_v2.iloc[i]
        ref = df_ref.iloc[i]
        cd_v1 = contact_details_v1[i]
        cd_v2 = contact_details_v2[i]
        ad_v1 = account_details_v1[i]
        ad_v2 = account_details_v2[i]
        
        # Bezpieczne pobieranie wartosci
        def safe_get(df_row, col, default=""):
            val = df_row.get(col, default)
            return val if pd.notna(val) else ""
        
        values = [
            i + 1,
            # Dane wejsciowe
            safe_get(v1, "input_raw_name"), safe_get(v1, "input_first_name"), safe_get(v1, "input_last_name"),
            safe_get(v1, "input_email"), safe_get(v1, "input_phone"), safe_get(v1, "input_company"),
            # V1 - normalizacja
            safe_get(v1, "norm_first_name"), safe_get(v1, "norm_last_name"), safe_get(v1, "norm_email"),
            safe_get(v1, "norm_phone"), safe_get(v1, "norm_company_name"), safe_get(v1, "norm_nip"), safe_get(v1, "norm_nip_valid"),
            # V1 - weryfikacja
            safe_get(v1, "contact_exists"), safe_get(v1, "contact_id"), safe_get(v1, "contact_name"),
            safe_get(v1, "contact_tier"), safe_get(v1, "contact_signals"),
            safe_get(v1, "account_exists"), safe_get(v1, "account_id"), safe_get(v1, "account_name"), safe_get(v1, "account_match"),
            safe_get(v1, "recommendation_action"), safe_get(v1, "recommendation_confidence"),
            # V2 - normalizacja
            safe_get(v2, "norm_first_name"), safe_get(v2, "norm_last_name"), safe_get(v2, "norm_email"),
            safe_get(v2, "norm_phone"), safe_get(v2, "norm_company_name"), safe_get(v2, "norm_nip"), safe_get(v2, "norm_nip_valid"),
            # V2 - weryfikacja
            safe_get(v2, "contact_exists"), safe_get(v2, "contact_id"), safe_get(v2, "contact_name"),
            safe_get(v2, "contact_tier"), safe_get(v2, "contact_signals"),
            safe_get(v2, "account_exists"), safe_get(v2, "account_id"), safe_get(v2, "account_name"), safe_get(v2, "account_match"),
            safe_get(v2, "recommendation_action"), safe_get(v2, "recommendation_confidence"),
            # Referencja
            cd_v1["ref_id"], cd_v1["ref_name"], ad_v1["ref_id"], ad_v1["ref_name"],
            # Wyniki
            cd_v1["result"], cd_v2["result"], ad_v1["result"], ad_v2["result"],
        ]
        
        for col, v in enumerate(values, 1):
            cell = ws_full.cell(row=row, column=col, value=v)
            cell.border = thin_border
            
            # Koloruj wyniki
            if v == "TRUE_POSITIVE" or v == "TRUE_NEGATIVE":
                cell.fill = winner_fill
            elif v == "FALSE_POSITIVE" or v == "FALSE_NEGATIVE":
                cell.fill = loser_fill
    
    # === ARKUSZ 3: DETAILS (skrocony) ===
    ws_details = wb.create_sheet("Details")
    
    detail_headers = [
        "Row", 
        "REF Contact ID", "REF Contact Name",
        "V1 Contact ID", "V1 Contact Name", "V1 Tier", "V1 Signals", "V1 Result",
        "V2 Contact ID", "V2 Contact Name", "V2 Tier", "V2 Signals", "V2 Result",
        "REF Account ID", "REF Account Name",
        "V1 Account ID", "V1 Account Name", "V1 Result",
        "V2 Account ID", "V2 Account Name", "V2 Result",
    ]
    
    for col, h in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for i in range(len(contact_details_v1)):
        row = i + 2
        cd_v1 = contact_details_v1[i]
        cd_v2 = contact_details_v2[i]
        ad_v1 = account_details_v1[i]
        ad_v2 = account_details_v2[i]
        
        values = [
            cd_v1["row"],
            cd_v1["ref_id"], cd_v1["ref_name"],
            cd_v1["test_id"], cd_v1["test_name"], cd_v1["test_tier"], cd_v1["test_signals"], cd_v1["result"],
            cd_v2["test_id"], cd_v2["test_name"], cd_v2["test_tier"], cd_v2["test_signals"], cd_v2["result"],
            ad_v1["ref_id"], ad_v1["ref_name"],
            ad_v1["test_id"], ad_v1["test_name"], ad_v1["result"],
            ad_v2["test_id"], ad_v2["test_name"], ad_v2["result"],
        ]
        
        for col, v in enumerate(values, 1):
            cell = ws_details.cell(row=row, column=col, value=v)
            cell.border = thin_border
            
            # Koloruj wyniki
            if v == "TRUE_POSITIVE" or v == "TRUE_NEGATIVE":
                cell.fill = winner_fill
            elif v == "FALSE_POSITIVE" or v == "FALSE_NEGATIVE":
                cell.fill = loser_fill
    
    # === ARKUSZ 4: MISMATCHES ===
    ws_mismatches = wb.create_sheet("Mismatches")
    
    mismatch_headers = ["Row", "Type", "Variant", "REF ID", "REF Name", "TEST ID", "TEST Name", "Result"]
    for col, h in enumerate(mismatch_headers, 1):
        cell = ws_mismatches.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    mismatch_row = 2
    
    for i in range(len(contact_details_v1)):
        cd_v1 = contact_details_v1[i]
        cd_v2 = contact_details_v2[i]
        ad_v1 = account_details_v1[i]
        ad_v2 = account_details_v2[i]
        
        # V1 Contact mismatch
        if cd_v1["result"] in ["FALSE_POSITIVE", "FALSE_NEGATIVE"]:
            values = [cd_v1["row"], "Contact", "V1", cd_v1["ref_id"], cd_v1["ref_name"], cd_v1["test_id"], cd_v1["test_name"], cd_v1["result"]]
            for col, v in enumerate(values, 1):
                cell = ws_mismatches.cell(row=mismatch_row, column=col, value=v)
                cell.border = thin_border
                if v in ["FALSE_POSITIVE", "FALSE_NEGATIVE"]:
                    cell.fill = loser_fill
            mismatch_row += 1
        
        # V2 Contact mismatch
        if cd_v2["result"] in ["FALSE_POSITIVE", "FALSE_NEGATIVE"]:
            values = [cd_v2["row"], "Contact", "V2", cd_v1["ref_id"], cd_v1["ref_name"], cd_v2["test_id"], cd_v2["test_name"], cd_v2["result"]]
            for col, v in enumerate(values, 1):
                cell = ws_mismatches.cell(row=mismatch_row, column=col, value=v)
                cell.border = thin_border
                if v in ["FALSE_POSITIVE", "FALSE_NEGATIVE"]:
                    cell.fill = loser_fill
            mismatch_row += 1
        
        # V1 Account mismatch
        if ad_v1["result"] in ["FALSE_POSITIVE", "FALSE_NEGATIVE"]:
            values = [ad_v1["row"], "Account", "V1", ad_v1["ref_id"], ad_v1["ref_name"], ad_v1["test_id"], ad_v1["test_name"], ad_v1["result"]]
            for col, v in enumerate(values, 1):
                cell = ws_mismatches.cell(row=mismatch_row, column=col, value=v)
                cell.border = thin_border
                if v in ["FALSE_POSITIVE", "FALSE_NEGATIVE"]:
                    cell.fill = loser_fill
            mismatch_row += 1
        
        # V2 Account mismatch
        if ad_v2["result"] in ["FALSE_POSITIVE", "FALSE_NEGATIVE"]:
            values = [ad_v2["row"], "Account", "V2", ad_v1["ref_id"], ad_v1["ref_name"], ad_v2["test_id"], ad_v2["test_name"], ad_v2["result"]]
            for col, v in enumerate(values, 1):
                cell = ws_mismatches.cell(row=mismatch_row, column=col, value=v)
                cell.border = thin_border
                if v in ["FALSE_POSITIVE", "FALSE_NEGATIVE"]:
                    cell.fill = loser_fill
            mismatch_row += 1
    
    # Zapisz
    wb.save(output_file)
    print(f"OK: Zapisano raport")


def print_summary(metrics):
    """Wyswietla podsumowanie na konsoli."""
    print("\n" + "=" * 80)
    print("PODSUMOWANIE POROWNANIA")
    print("=" * 80)
    
    print("\n--- KONTAKTY ---")
    print(f"                    V1 (STRUCT)    V2 (RAW)")
    print(f"  True Positive:    {metrics['contact']['v1']['true_positive']:>6}         {metrics['contact']['v2']['true_positive']:>6}")
    print(f"  False Negative:   {metrics['contact']['v1']['false_negative']:>6}         {metrics['contact']['v2']['false_negative']:>6}")
    print(f"  False Positive:   {metrics['contact']['v1']['false_positive']:>6}         {metrics['contact']['v2']['false_positive']:>6}")
    print(f"  Tier 4:           {metrics['contact']['v1']['tier_4']:>6}         {metrics['contact']['v2']['tier_4']:>6}")
    print(f"  Tier 3:           {metrics['contact']['v1']['tier_3']:>6}         {metrics['contact']['v2']['tier_3']:>6}")
    
    print("\n--- FIRMY ---")
    print(f"                    V1 (STRUCT)    V2 (RAW)")
    print(f"  True Positive:    {metrics['account']['v1']['true_positive']:>6}         {metrics['account']['v2']['true_positive']:>6}")
    print(f"  False Negative:   {metrics['account']['v1']['false_negative']:>6}         {metrics['account']['v2']['false_negative']:>6}")
    print(f"  False Positive:   {metrics['account']['v1']['false_positive']:>6}         {metrics['account']['v2']['false_positive']:>6}")
    
    print("\n--- NORMALIZACJA ---")
    print(f"                    V1 (STRUCT)    V2 (RAW)")
    print(f"  Imie:             {metrics['normalization']['v1']['first_name_filled']:>6}         {metrics['normalization']['v2']['first_name_filled']:>6}")
    print(f"  Nazwisko:         {metrics['normalization']['v1']['last_name_filled']:>6}         {metrics['normalization']['v2']['last_name_filled']:>6}")
    print(f"  Email:            {metrics['normalization']['v1']['email_filled']:>6}         {metrics['normalization']['v2']['email_filled']:>6}")
    print(f"  Firma:            {metrics['normalization']['v1']['company_filled']:>6}         {metrics['normalization']['v2']['company_filled']:>6}")
    print(f"  NIP znaleziony:   {metrics['normalization']['v1']['nip_found']:>6}         {metrics['normalization']['v2']['nip_found']:>6}")
    
    print("\n--- REKOMENDACJE ---")
    print(f"                    V1 (STRUCT)    V2 (RAW)")
    print(f"  link_to_existing: {metrics['recommendation']['v1']['link_to_existing']:>6}         {metrics['recommendation']['v2']['link_to_existing']:>6}")
    print(f"  create_new:       {metrics['recommendation']['v1']['create_new']:>6}         {metrics['recommendation']['v2']['create_new']:>6}")
    print(f"  review_required:  {metrics['recommendation']['v1']['review_required']:>6}         {metrics['recommendation']['v2']['review_required']:>6}")
    
    # Werdykt
    v1_wins = 0
    v2_wins = 0
    
    # Kontakty - TP wieksze, FN mniejsze, FP mniejsze
    if metrics['contact']['v1']['true_positive'] > metrics['contact']['v2']['true_positive']:
        v1_wins += 1
    elif metrics['contact']['v2']['true_positive'] > metrics['contact']['v1']['true_positive']:
        v2_wins += 1
    
    if metrics['contact']['v1']['false_negative'] < metrics['contact']['v2']['false_negative']:
        v1_wins += 1
    elif metrics['contact']['v2']['false_negative'] < metrics['contact']['v1']['false_negative']:
        v2_wins += 1
    
    # Firmy
    if metrics['account']['v1']['true_positive'] > metrics['account']['v2']['true_positive']:
        v1_wins += 1
    elif metrics['account']['v2']['true_positive'] > metrics['account']['v1']['true_positive']:
        v2_wins += 1
    
    # Rekomendacje - link_to_existing wieksze
    if metrics['recommendation']['v1']['link_to_existing'] > metrics['recommendation']['v2']['link_to_existing']:
        v1_wins += 1
    elif metrics['recommendation']['v2']['link_to_existing'] > metrics['recommendation']['v1']['link_to_existing']:
        v2_wins += 1
    
    print("\n" + "=" * 80)
    print("WERDYKT")
    print("=" * 80)
    print(f"  Wariant 1 (STRUKTURALNY): {v1_wins} punktow")
    print(f"  Wariant 2 (SUROWY TEKST): {v2_wins} punktow")
    
    if v1_wins > v2_wins:
        print("\n  >>> ZWYCIEZCA: WARIANT 1 (STRUKTURALNY) <<<")
    elif v2_wins > v1_wins:
        print("\n  >>> ZWYCIEZCA: WARIANT 2 (SUROWY TEKST) <<<")
    else:
        print("\n  >>> REMIS <<<")
    
    print("=" * 80)


def main():
    """Glowna funkcja."""
    print("=" * 80)
    print("POROWNANIE WARIANTOW AI")
    print("Strukturalny vs Surowy tekst")
    print("=" * 80)
    
    # Wczytaj dane
    df_ref, df_v1, df_v2 = load_data()
    ref_cols = get_ref_columns(df_ref)
    
    # Oblicz metryki
    metrics = create_comparison_report(df_ref, df_v1, df_v2, ref_cols)
    
    # Wyswietl podsumowanie
    print_summary(metrics)
    
    # Zapisz raport
    write_excel_report(metrics, df_ref, df_v1, df_v2, ref_cols, "test_comparison_report.xlsx")
    
    print(f"\nRaport zapisany w: test_comparison_report.xlsx")


if __name__ == "__main__":
    main()
