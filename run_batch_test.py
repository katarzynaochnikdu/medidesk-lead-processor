"""
Skrypt do uruchamiania testow paczkami z pelna dokumentacja.
Kazda paczka generuje szczegolowy raport z transformacjami danych.
"""

import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

# Wylacz buforowanie stdout
sys.stdout.reconfigure(line_buffering=True)

import pandas as pd

# Dodaj src do path
sys.path.insert(0, str(Path(__file__).parent))

from src.config import get_settings
from src.services.data_normalizer import DataNormalizerService


class BatchTestRunner:
    """Runner do testow paczkami z pelna dokumentacja."""
    
    def __init__(
        self,
        test_file: str,
        reference_file: str,
        output_dir: str = "test_cycles",
        batch_size: int = 100,
    ):
        self.test_file = test_file
        self.reference_file = reference_file
        self.output_dir = Path(output_dir)
        self.batch_size = batch_size
        
        # Wczytaj dane
        self.df_test = pd.read_excel(test_file)
        self.df_ref = pd.read_excel(reference_file)
        
        # Mapuj kolumny referencji
        self.ref_cols = {
            "contact_id": self.df_ref.columns[7],
            "contact_name": self.df_ref.columns[8],
            "account_id": self.df_ref.columns[9],
            "account_name": self.df_ref.columns[10],
        }
        
        # Utworz katalog wyjsciowy
        self.output_dir.mkdir(exist_ok=True)
        
        print(f"=== BATCH TEST RUNNER ===")
        print(f"Test file: {test_file} ({len(self.df_test)} rows)")
        print(f"Reference file: {reference_file} ({len(self.df_ref)} rows)")
        print(f"Batch size: {batch_size}")
        print(f"Output dir: {output_dir}")
    
    def get_batch_ranges(self) -> list[tuple[int, int, str]]:
        """Zwraca zakresy paczek."""
        total = len(self.df_test)
        batches = []
        
        start = 0
        batch_num = 1
        while start < total:
            end = min(start + self.batch_size, total)
            batches.append((start, end, f"batch_{batch_num}"))
            start = end
            batch_num += 1
        
        return batches
    
    async def run_batch(
        self,
        start_idx: int,
        end_idx: int,
        batch_name: str,
        variant: str = "structural",
    ) -> pd.DataFrame:
        """
        Uruchamia test na paczce danych.
        
        Args:
            start_idx: Indeks poczatkowy
            end_idx: Indeks koncowy
            batch_name: Nazwa paczki
            variant: "structural" lub "raw_text"
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        print(f"\n{'='*80}")
        print(f"BATCH: {batch_name} | VARIANT: {variant}")
        print(f"Rows: {start_idx+1} - {end_idx} ({end_idx - start_idx} records)")
        print(f"Timestamp: {timestamp}")
        print(f"{'='*80}\n")
        
        # Inicjalizuj serwis
        settings = get_settings()
        service = DataNormalizerService(settings=settings, use_mocks=False)
        
        results = []
        
        for idx in range(start_idx, end_idx):
            row = self.df_test.iloc[idx]
            ref_row = self.df_ref.iloc[idx]
            
            # Przygotuj dane wejsciowe
            if variant == "structural":
                lead_data = self._prepare_structural_data(row)
            else:
                lead_data = self._prepare_raw_text_data(row)
            
            display_name = lead_data.get("raw_name", "")[:40] or f"{lead_data.get('first_name', '')} {lead_data.get('last_name', '')}".strip()
            print(f"[{idx+1}/{end_idx}] {display_name}")
            
            try:
                # Przetwarzanie
                output = await service.process_lead(
                    raw_data=lead_data,
                    skip_ai=False,
                    skip_gus=False,
                    skip_duplicates=False,
                )
                
                # Zbierz wyniki
                result = self._collect_result(
                    idx=idx,
                    row=row,
                    ref_row=ref_row,
                    lead_data=lead_data,
                    output=output,
                    variant=variant,
                )
                
                print(f"  -> {result['norm_first_name']} {result['norm_last_name']} | Contact: {result['contact_exists']} | Account: {result['account_exists']}")
                
            except Exception as e:
                print(f"  ERROR: {e}")
                result = self._collect_error_result(idx, row, ref_row, lead_data, str(e), variant)
            
            results.append(result)
        
        # Zamknij polaczenia
        await service.close()
        
        # Zapisz wyniki
        df_results = pd.DataFrame(results)
        output_file = self.output_dir / f"{batch_name}_{variant}_{timestamp}.xlsx"
        
        self._save_results(df_results, output_file)
        
        print(f"\n==> Zapisano: {output_file}")
        
        return df_results
    
    def _prepare_structural_data(self, row: pd.Series) -> dict:
        """Przygotowuje dane strukturalne."""
        phone = row.get("Telefon")
        if pd.notna(phone):
            phone = str(int(phone)) if isinstance(phone, (int, float)) else str(phone)
        else:
            phone = None
        
        data = {
            "id": row.get("Id rekordu"),
            "raw_name": row.get("Marketing Lead - nazwa"),
            "email": row.get("Adres Email"),
            "phone": phone,
            "company": row.get("Firma"),
            "first_name": row.get("Imie") if pd.notna(row.get("Imie")) else row.get("Imię"),
            "last_name": row.get("Nazwisko"),
        }
        
        return {k: (str(v) if pd.notna(v) and v is not None else None) for k, v in data.items()}
    
    def _prepare_raw_text_data(self, row: pd.Series) -> dict:
        """Przygotowuje dane jako surowy tekst."""
        parts = []
        
        if pd.notna(row.get("Marketing Lead - nazwa")):
            parts.append(f"Marketing Lead: {row['Marketing Lead - nazwa']}")
        if pd.notna(row.get("Firma")):
            parts.append(f"Firma: {row['Firma']}")
        
        first_name = row.get("Imie") if pd.notna(row.get("Imie")) else row.get("Imię")
        if pd.notna(first_name):
            parts.append(f"Imie: {first_name}")
        if pd.notna(row.get("Nazwisko")):
            parts.append(f"Nazwisko: {row['Nazwisko']}")
        if pd.notna(row.get("Adres Email")):
            parts.append(f"Email: {row['Adres Email']}")
        if pd.notna(row.get("Telefon")):
            phone = row["Telefon"]
            if isinstance(phone, (int, float)):
                phone = str(int(phone))
            parts.append(f"Telefon: {phone}")
        
        raw_text = "\n".join(parts)
        
        return {
            "id": row.get("Id rekordu"),
            "raw_name": raw_text,
            "email": None,
            "phone": None,
            "company": None,
            "first_name": None,
            "last_name": None,
        }
    
    def _collect_result(
        self,
        idx: int,
        row: pd.Series,
        ref_row: pd.Series,
        lead_data: dict,
        output,
        variant: str,
    ) -> dict:
        """Zbiera pelny wynik z transformacjami."""
        normalized = output.normalized
        gus_data = output.gus_data
        duplicates = output.duplicates
        recommendation = output.recommendation
        warnings = output.warnings
        
        # Kontakt
        contact_result = duplicates.contact
        contact_candidates = contact_result.candidates
        
        contact_name = ""
        contact_tier = 0
        contact_signals = ""
        
        if contact_candidates:
            best = contact_candidates[0]
            contact_name = best.name
            contact_tier = best.tier
            signals = best.signals
            signal_parts = []
            if signals.E: signal_parts.append("E")
            if signals.P: signal_parts.append("P")
            if signals.L: signal_parts.append("L")
            if signals.F: signal_parts.append("F")
            if signals.A: signal_parts.append("A")
            contact_signals = "+".join(signal_parts)
        
        # Firma
        account_result = duplicates.account
        account_candidates = account_result.candidates
        
        account_name = ""
        account_match = ""
        if account_candidates:
            best_acc = account_candidates[0]
            account_name = best_acc.name
            account_match = best_acc.match_reason
        
        # NIP source
        nip_source = ""
        if normalized.nip:
            if any("Brave Search" in w for w in warnings):
                nip_source = "Brave"
            else:
                nip_source = "AI/ekstrakcja"
        
        # Referencja
        ref_contact_id = ref_row[self.ref_cols["contact_id"]]
        ref_contact_name = ref_row[self.ref_cols["contact_name"]]
        ref_account_id = ref_row[self.ref_cols["account_id"]]
        ref_account_name = ref_row[self.ref_cols["account_name"]]
        
        ref_contact_exists = pd.notna(ref_contact_id) and str(ref_contact_id).lower() != "nowy"
        ref_account_exists = pd.notna(ref_account_id)
        
        test_contact_exists = contact_result.exists
        test_account_exists = account_result.exists
        
        # Normalizuj ID do porownania (usun prefix zcrm_, konwertuj float na int string)
        def normalize_id(id_val):
            if pd.isna(id_val) or id_val is None:
                return None
            s = str(id_val)
            # Usun prefix zcrm_
            if s.startswith("zcrm_"):
                s = s[5:]
            # Konwertuj notacje naukowa do int
            try:
                if 'e' in s.lower() or '.' in s:
                    s = str(int(float(s)))
            except:
                pass
            return s
        
        ref_contact_id_norm = normalize_id(ref_contact_id)
        test_contact_id_norm = normalize_id(contact_result.primary_id)
        ref_account_id_norm = normalize_id(ref_account_id)
        test_account_id_norm = normalize_id(account_result.parent_id)
        
        # Wynik porownania
        if ref_contact_exists and test_contact_exists:
            if ref_contact_id_norm == test_contact_id_norm:
                contact_match_result = "TRUE_POSITIVE"
            else:
                contact_match_result = "FALSE_POSITIVE"
        elif ref_contact_exists and not test_contact_exists:
            contact_match_result = "FALSE_NEGATIVE"
        elif not ref_contact_exists and test_contact_exists:
            contact_match_result = "FALSE_POSITIVE"
        else:
            contact_match_result = "TRUE_NEGATIVE"
        
        if ref_account_exists and test_account_exists:
            if ref_account_id_norm == test_account_id_norm:
                account_match_result = "TRUE_POSITIVE"
            else:
                account_match_result = "FALSE_POSITIVE"
        elif ref_account_exists and not test_account_exists:
            account_match_result = "FALSE_NEGATIVE"
        elif not ref_account_exists and test_account_exists:
            account_match_result = "FALSE_POSITIVE"
        else:
            account_match_result = "TRUE_NEGATIVE"
        
        return {
            "row_num": idx + 1,
            "status": "OK",
            "variant": variant.upper(),
            
            # Input
            "input_id": lead_data.get("id", ""),
            "input_raw_name": lead_data.get("raw_name", "")[:100] if lead_data.get("raw_name") else "",
            "input_first_name": lead_data.get("first_name", ""),
            "input_last_name": lead_data.get("last_name", ""),
            "input_email": lead_data.get("email", ""),
            "input_phone": lead_data.get("phone", ""),
            "input_company": lead_data.get("company", ""),
            
            # Normalizacja AI
            "norm_first_name": normalized.first_name or "",
            "norm_last_name": normalized.last_name or "",
            "norm_email": normalized.email or "",
            "norm_phone": normalized.phone_formatted or "",
            "norm_company_name": normalized.company_name or "",
            "norm_company_rejected": "TAK" if normalized.company_rejected else "NIE",
            "norm_company_rejected_reason": normalized.company_rejected_reason or "",
            "norm_nip": normalized.nip_formatted or "",
            "norm_nip_valid": "TAK" if normalized.nip_valid else "NIE",
            "norm_nip_source": nip_source,
            
            # GUS
            "gus_found": "TAK" if gus_data.found else "NIE",
            "gus_name": gus_data.full_name or "",
            
            # Kontakt - test
            "contact_exists": "TAK" if test_contact_exists else "NIE",
            "contact_id": test_contact_id_norm or "",
            "contact_name": contact_name,
            "contact_tier": contact_tier,
            "contact_signals": contact_signals,
            "contact_candidates_count": len(contact_candidates),
            "contact_needs_review": "TAK" if contact_result.needs_review else "NIE",
            
            # Firma - test
            "account_exists": "TAK" if test_account_exists else "NIE",
            "account_id": test_account_id_norm or "",
            "account_name": account_name,
            "account_match": account_match,
            "account_candidates_count": len(account_candidates),
            
            # Referencja (ID znormalizowane)
            "ref_contact_id": ref_contact_id_norm if ref_contact_exists else "NOWY",
            "ref_contact_name": ref_contact_name if pd.notna(ref_contact_name) else "",
            "ref_account_id": ref_account_id_norm if ref_account_exists else "",
            "ref_account_name": ref_account_name if pd.notna(ref_account_name) else "",
            
            # Wynik porownania
            "contact_match_result": contact_match_result,
            "account_match_result": account_match_result,
            
            # Rekomendacja
            "recommendation_action": recommendation.action,
            "recommendation_confidence": recommendation.confidence,
            "recommendation_reason": recommendation.reason[:100] if recommendation.reason else "",
            
            # Ostrzezenia
            "warnings": "; ".join(warnings)[:200] if warnings else "",
        }
    
    def _collect_error_result(
        self,
        idx: int,
        row: pd.Series,
        ref_row: pd.Series,
        lead_data: dict,
        error: str,
        variant: str,
    ) -> dict:
        """Zbiera wynik bledu."""
        return {
            "row_num": idx + 1,
            "status": "ERROR",
            "variant": variant.upper(),
            "input_id": lead_data.get("id", ""),
            "input_raw_name": lead_data.get("raw_name", "")[:100] if lead_data.get("raw_name") else "",
            "input_first_name": lead_data.get("first_name", ""),
            "input_last_name": lead_data.get("last_name", ""),
            "input_email": lead_data.get("email", ""),
            "input_phone": lead_data.get("phone", ""),
            "input_company": lead_data.get("company", ""),
            "error": error[:200],
        }
    
    def _save_results(self, df: pd.DataFrame, output_file: Path):
        """Zapisuje wyniki do Excel z formatowaniem."""
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Results', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Results']
            
            # Style
            header_font = Font(bold=True)
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Naglowki
            for cell in worksheet[1]:
                cell.font = header_font
            
            # Koloruj wyniki
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                for cell in row:
                    if cell.value == "TRUE_POSITIVE" or cell.value == "TRUE_NEGATIVE":
                        cell.fill = green_fill
                    elif cell.value == "FALSE_POSITIVE" or cell.value == "FALSE_NEGATIVE":
                        cell.fill = red_fill
            
            # Auto-width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def analyze_batch(self, df_v1: pd.DataFrame, df_v2: pd.DataFrame, batch_name: str) -> dict:
        """Analizuje wyniki paczki i generuje metryki."""
        
        def calc_metrics(df: pd.DataFrame, variant: str) -> dict:
            total = len(df)
            ok = (df["status"] == "OK").sum()
            
            # Kontakty
            contact_tp = (df["contact_match_result"] == "TRUE_POSITIVE").sum()
            contact_tn = (df["contact_match_result"] == "TRUE_NEGATIVE").sum()
            contact_fp = (df["contact_match_result"] == "FALSE_POSITIVE").sum()
            contact_fn = (df["contact_match_result"] == "FALSE_NEGATIVE").sum()
            
            # Firmy
            account_tp = (df["account_match_result"] == "TRUE_POSITIVE").sum()
            account_tn = (df["account_match_result"] == "TRUE_NEGATIVE").sum()
            account_fp = (df["account_match_result"] == "FALSE_POSITIVE").sum()
            account_fn = (df["account_match_result"] == "FALSE_NEGATIVE").sum()
            
            # Tier distribution
            tier_4 = (df["contact_tier"] == 4).sum()
            tier_3 = (df["contact_tier"] == 3).sum()
            tier_2 = (df["contact_tier"] == 2).sum()
            tier_0 = (df["contact_tier"] == 0).sum()
            
            # Rekomendacje
            link = (df["recommendation_action"] == "link_to_existing").sum()
            create = (df["recommendation_action"] == "create_new").sum()
            review = (df["recommendation_action"] == "review_required").sum()
            
            return {
                "variant": variant,
                "total": total,
                "ok": ok,
                "errors": total - ok,
                "contact_tp": contact_tp,
                "contact_tn": contact_tn,
                "contact_fp": contact_fp,
                "contact_fn": contact_fn,
                "contact_precision": contact_tp / (contact_tp + contact_fp) if (contact_tp + contact_fp) > 0 else 0,
                "contact_recall": contact_tp / (contact_tp + contact_fn) if (contact_tp + contact_fn) > 0 else 0,
                "account_tp": account_tp,
                "account_tn": account_tn,
                "account_fp": account_fp,
                "account_fn": account_fn,
                "account_precision": account_tp / (account_tp + account_fp) if (account_tp + account_fp) > 0 else 0,
                "account_recall": account_tp / (account_tp + account_fn) if (account_tp + account_fn) > 0 else 0,
                "tier_4": tier_4,
                "tier_3": tier_3,
                "tier_2": tier_2,
                "tier_0": tier_0,
                "rec_link": link,
                "rec_create": create,
                "rec_review": review,
            }
        
        metrics_v1 = calc_metrics(df_v1, "V1_STRUCTURAL")
        metrics_v2 = calc_metrics(df_v2, "V2_RAW_TEXT")
        
        return {
            "batch_name": batch_name,
            "timestamp": datetime.now().isoformat(),
            "v1": metrics_v1,
            "v2": metrics_v2,
        }
    
    def generate_analysis_report(self, metrics: dict, batch_name: str) -> str:
        """Generuje raport analizy w formacie Markdown."""
        v1 = metrics["v1"]
        v2 = metrics["v2"]
        
        report = f"""# Analiza {batch_name}

## Data: {metrics['timestamp']}

## Podsumowanie

| Metryka | V1 (Structural) | V2 (Raw Text) | Lepszy |
|---------|-----------------|---------------|--------|
| Przetworzonych | {v1['ok']}/{v1['total']} | {v2['ok']}/{v2['total']} | - |
| Bledow | {v1['errors']} | {v2['errors']} | {'V1' if v1['errors'] < v2['errors'] else 'V2' if v2['errors'] < v1['errors'] else 'REMIS'} |

## Kontakty

| Metryka | V1 | V2 | Lepszy |
|---------|----|----|--------|
| True Positive | {v1['contact_tp']} | {v2['contact_tp']} | {'V1' if v1['contact_tp'] > v2['contact_tp'] else 'V2' if v2['contact_tp'] > v1['contact_tp'] else 'REMIS'} |
| True Negative | {v1['contact_tn']} | {v2['contact_tn']} | - |
| False Positive | {v1['contact_fp']} | {v2['contact_fp']} | {'V1' if v1['contact_fp'] < v2['contact_fp'] else 'V2' if v2['contact_fp'] < v1['contact_fp'] else 'REMIS'} |
| False Negative | {v1['contact_fn']} | {v2['contact_fn']} | {'V1' if v1['contact_fn'] < v2['contact_fn'] else 'V2' if v2['contact_fn'] < v1['contact_fn'] else 'REMIS'} |
| Precision | {v1['contact_precision']:.2%} | {v2['contact_precision']:.2%} | {'V1' if v1['contact_precision'] > v2['contact_precision'] else 'V2' if v2['contact_precision'] > v1['contact_precision'] else 'REMIS'} |
| Recall | {v1['contact_recall']:.2%} | {v2['contact_recall']:.2%} | {'V1' if v1['contact_recall'] > v2['contact_recall'] else 'V2' if v2['contact_recall'] > v1['contact_recall'] else 'REMIS'} |

## Tier Distribution

| Tier | V1 | V2 |
|------|----|----|
| Tier 4 | {v1['tier_4']} | {v2['tier_4']} |
| Tier 3 | {v1['tier_3']} | {v2['tier_3']} |
| Tier 2 | {v1['tier_2']} | {v2['tier_2']} |
| Tier 0 | {v1['tier_0']} | {v2['tier_0']} |

## Firmy

| Metryka | V1 | V2 | Lepszy |
|---------|----|----|--------|
| True Positive | {v1['account_tp']} | {v2['account_tp']} | {'V1' if v1['account_tp'] > v2['account_tp'] else 'V2' if v2['account_tp'] > v1['account_tp'] else 'REMIS'} |
| False Positive | {v1['account_fp']} | {v2['account_fp']} | {'V1' if v1['account_fp'] < v2['account_fp'] else 'V2' if v2['account_fp'] < v1['account_fp'] else 'REMIS'} |
| False Negative | {v1['account_fn']} | {v2['account_fn']} | {'V1' if v1['account_fn'] < v2['account_fn'] else 'V2' if v2['account_fn'] < v1['account_fn'] else 'REMIS'} |
| Precision | {v1['account_precision']:.2%} | {v2['account_precision']:.2%} | - |
| Recall | {v1['account_recall']:.2%} | {v2['account_recall']:.2%} | - |

## Rekomendacje

| Akcja | V1 | V2 |
|-------|----|----|
| link_to_existing | {v1['rec_link']} | {v2['rec_link']} |
| create_new | {v1['rec_create']} | {v2['rec_create']} |
| review_required | {v1['rec_review']} | {v2['rec_review']} |

## Wnioski

TODO: Dodac wnioski po analizie

## Proponowane poprawki

TODO: Dodac propozycje poprawek

"""
        return report


async def run_batch(batch_num: int = 1):
    """Uruchamia wybrany batch."""
    runner = BatchTestRunner(
        test_file="test set 2.xlsx",
        reference_file="test set reference 2.xlsx",
        output_dir="test_cycles",
        batch_size=100,
    )
    
    batches = runner.get_batch_ranges()
    print(f"\nZdefiniowane paczki: {batches}")
    
    if batch_num < 1 or batch_num > len(batches):
        print(f"ERROR: Batch {batch_num} nie istnieje. Dostepne: 1-{len(batches)}")
        return None
    
    # Wybrany batch
    start, end, name = batches[batch_num - 1]
    
    print(f"\n{'#'*80}")
    print(f"# BATCH {batch_num}: wiersze {start+1}-{end}")
    print(f"{'#'*80}")
    
    # V1 - Structural
    df_v1 = await runner.run_batch(start, end, name, variant="structural")
    
    # V2 - Raw text
    df_v2 = await runner.run_batch(start, end, name, variant="raw_text")
    
    # Analiza
    metrics = runner.analyze_batch(df_v1, df_v2, name)
    
    # Raport
    report = runner.generate_analysis_report(metrics, name)
    
    report_file = runner.output_dir / f"{name}_analysis.md"
    with open(report_file, "w", encoding="utf-8") as f:
        f.write(report)
    
    print(f"\n==> Raport analizy: {report_file}")
    
    # Zapisz metryki jako JSON
    metrics_file = runner.output_dir / f"{name}_metrics.json"
    with open(metrics_file, "w", encoding="utf-8") as f:
        json.dump(metrics, f, indent=2, default=str)
    
    print(f"==> Metryki JSON: {metrics_file}")
    
    # Wyswietl podsumowanie
    print(f"\n{'='*80}")
    print(f"PODSUMOWANIE BATCH {batch_num}")
    print(f"{'='*80}")
    print(f"\nKontakty:")
    print(f"  V1: TP={metrics['v1']['contact_tp']}, FN={metrics['v1']['contact_fn']}, FP={metrics['v1']['contact_fp']}")
    print(f"  V2: TP={metrics['v2']['contact_tp']}, FN={metrics['v2']['contact_fn']}, FP={metrics['v2']['contact_fp']}")
    print(f"\nFirmy:")
    print(f"  V1: TP={metrics['v1']['account_tp']}, FN={metrics['v1']['account_fn']}, FP={metrics['v1']['account_fp']}")
    print(f"  V2: TP={metrics['v2']['account_tp']}, FN={metrics['v2']['account_fn']}, FP={metrics['v2']['account_fp']}")
    
    return metrics


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Run batch test")
    parser.add_argument("--batch", type=int, default=1, help="Batch number (1, 2, or 3)")
    args = parser.parse_args()
    asyncio.run(run_batch(args.batch))
